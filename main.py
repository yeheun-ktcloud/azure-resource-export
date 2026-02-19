# -*- coding: utf-8 -*-

import os
import requests
import textwrap
from azure.mgmt.resource import SubscriptionClient
from datetime import datetime
from typing import List, Dict, Any, Tuple, Set
from openpyxl import load_workbook

from azure.identity import DefaultAzureCredential
from azure.mgmt.resourcegraph import ResourceGraphClient
from azure.mgmt.resourcegraph.models import QueryRequest
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.containerservice import ContainerServiceClient

from openpyxl import Workbook


# =========================
# 0. 사용자 입력 구간
# =========================

print("➡ Azure 인증: DefaultAzureCredential 방식 (Azure CLI/Managed Identity 등) 사용")

def ask_subscription(label: str) -> str:
    return input(f"{label} Subscription ID? (없으면 엔터): ").strip()

def ask_resource_groups(label: str) -> List[str]:
    raw = input(f"{label} Resource Group Name? (여러 개면 콤마로 구분, 없으면 엔터): ").strip()
    if not raw:
        return []
    # trim + lower 기준으로 중복 제거(원본 표기는 첫 등장 유지)
    seen = set()
    out = []
    for x in [t.strip() for t in raw.split(",") if t.strip()]:
        key = x.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(x)
    return out

DEV_SUB = ask_subscription("DEV")
STG_SUB = ask_subscription("STG")
PRD_SUB = ask_subscription("PRD")

DEV_RGS = ask_resource_groups("DEV")
STG_RGS = ask_resource_groups("STG")
PRD_RGS = ask_resource_groups("PRD")

SUB_ENV_MAP: Dict[str, str] = {k: v for k, v in {"DEV": DEV_SUB, "STG": STG_SUB, "PRD": PRD_SUB}.items() if v}
RG_ENV_MAP: Dict[str, List[str]] = {k: v for k, v in {"DEV": DEV_RGS, "STG": STG_RGS, "PRD": PRD_RGS}.items() if v}

if not SUB_ENV_MAP:
    raise RuntimeError("조회할 Azure 구독이 없습니다. DEV/STG/PRD 중 최소 1개는 입력해야 합니다.")

if not RG_ENV_MAP:
    raise RuntimeError("조회할 Resource Group이 없습니다. DEV/STG/PRD 중 최소 1개 이상 입력해야 합니다.")


# ✅ 단위서비스 정보 입력
service_name = input("단위서비스코드? : ").strip()
service_abbr = input("단위서비스 약어? : ").strip()

if not service_name or not service_abbr:
    raise RuntimeError("단위서비스명과 단위서비스 약어는 필수 입력입니다.")

out_dir = input("출력 폴더 경로? (없으면 엔터=현재 폴더): ").strip() or "."
os.makedirs(out_dir, exist_ok=True)



# =========================
# 1-1. Azure 공통
# =========================

def get_cred():
    return DefaultAzureCredential()

def get_arg_client():
    return ResourceGraphClient(get_cred())

def get_arm_token() -> str:
    return get_cred().get_token("https://management.azure.com/.default").token

def normalize_arg_query(q: str) -> str:
    q = textwrap.dedent(q).strip()
    # 흔한 유니코드 파이프/스마트쿼트 정리
    q = (q.replace("│", "|")
           .replace("｜", "|")
           .replace("’", "'")
           .replace("‘", "'")
           .replace("“", '"')
           .replace("”", '"'))
    return q

def esc_kql_str(s: str) -> str:
    # KQL 문자열 안에서 ' 를 '' 로 이스케이프
    return (s or "").replace("'", "''")

def add_rg_filter(kql: str, rg_name: str) -> str:
    rg = (rg_name or "").strip()
    if not rg:
        return kql
    rg_esc = esc_kql_str(rg)
    # Resources 바로 다음 줄에 where 추가하는 방식(가장 단순/명확)
    return kql.replace(
        "Resources",
        f"Resources\n| where resourceGroup =~ '{rg_esc}'",
        1
    )

def arm_get(url: str, api_version: str, timeout: int = 30) -> Dict[str, Any]:
    headers = {"Authorization": f"Bearer {get_arm_token()}"}
    r = requests.get(url, headers=headers, params={"api-version": api_version}, timeout=timeout)
    r.raise_for_status()
    return r.json() or {}

def get_public_ip_by_id(pip_id: str, cache: Dict[str, str]) -> str:
    if not pip_id:
        return ""
    pid = pip_id.lower()
    if pid in cache:
        return cache[pid]

    try:
        data = arm_get(f"https://management.azure.com{pip_id}", api_version="2023-09-01")
        ip = (data.get("properties") or {}).get("ipAddress", "") or ""
        cache[pid] = ip
        return ip
    except Exception:
        cache[pid] = ""
        return ""

# =========================
# 1-2. 구독 내 존재 리소스 타입 스캔
# =========================

def discover_resource_types(subscription_id: str, rg_name: str) -> Set[str]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| summarize cnt=count() by type
| where cnt > 0
| project type
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=[subscription_id], query=query)
    result = client.resources(req)
    return {str(r["type"]).lower() for r in result.data}


def should_run(types_in_sub: Set[str], expected_types: List[str]) -> bool:
    """
    expected_types 중 하나라도 구독 내에 존재하면 True
    """
    exp = [t.lower() for t in expected_types if t]
    return any(t in types_in_sub for t in exp)


# =========================
# 1-3. Sheet Headers
# =========================

summary_headers = [
    "Subscription Name", "ResourceGroup", "ResourceType", "ResourceName", "IP/FQDN"
]
appgw_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location","SKU",
    "MinCapacity","MaxCapacity","Public IP","Private IP"
]
lb_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "SKU","Tier","Frontend IP Name","Frontend IP","Backend Pool Name","Load Balancing Rule Name","Health Probe Name"
]
vm_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location","SKU","vCPU","Memory","OS",
    "Private IP","OS Disk","OS Disk Size(GiB)","Data Disk","Data Disk Size(GiB)"
]
vmss_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "SKU","vCPU","Memory","OS","OS Disk","OS Disk Size(GiB)","Data Disk","Data Disk Size(GiB)",
    "Orchestration Mode","Instance Count"
]
acr_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location","SKU","Login Server"
]
ca_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Container App Environment","WorkloadProfile","WorkloadProfile CPU","WorkloadProfile Memory","Application URL"
]
aks_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "OS","Nodepool","Node Count","Node CPU","Node Memory(GB)","OS Disk Size(GB)"
]
functionapp_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Default Domain","OS","App Service Plan","Pricing Tier","Runtime Version"
]
logicapp_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Default Domain","OS","App Service Plan","Pricing Tier","Runtime Version"
]
pg_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location","SKU",
    "Version","Storage Size","Storage IOPS","HA","Endpoint"
]
kv_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location","SKU","Vault URI"
]
st_sa_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Kind","Performance","Namespace","Access Tier","Redundancy"
]
file_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Storage Account Name","Storage Account Kind","Media Tier","Redundancy","Billing Model","Access Tier","Quota(GB)"
]
aisearch_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "SKU","Replica Count","Partition Count","URL"
]
docint_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Kind","SKU","Endpoint"
]
aml_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Container Registry","Key Vault","Application Insights"
]
eh_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Tier","Throughput Units","Auto Inflate","Max Throughput Units",
    "Premium Partitions","SKU Capacity"
]
openai_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location","SKU","Endpoint"
]
mysql_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type",
    "Location","SKU","Version","Storage Size","Storage IOPS","HA","Endpoint"
]
managed_redis_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Tier","Instance","HA","Endpoint"
]
docdb_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Cluster Tier","Shard Count","Storage Size(GB)","HA"
]
cosmos_mongo_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Write Regions","Write Region Mode","RU Throughput","URI"
]
cosmos_pg_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Tier","Compute Size"
]
sqlmi_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Service Tier","Hardware Generation","vCore","Memory(GB)","Storage Size(GB)","Zone Redundant","Host"
]
vnet_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "vNet IP Range","Subnet Name","Subnet IP Range","Subnet NSG","Subnet UDR"
]
private_endpoint_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Private IP","Connected Resource"
]
law_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "SKU","Retention(Days)","Public Network Access"
]
appinsights_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location",
    "Application Type","Workspace(ResourceId)"
]
managed_identity_headers = [
    "Environment","SubscriptionID","ResourceGroup","ResourceName","Type","Location"
]


# =========================
# 2. Summary
# =========================
def get_subscription_name_map(subscription_ids: List[str]) -> Dict[str, str]:
    """
    subscriptionId -> subscription displayName
    실패 시 subscriptionId 그대로 fallback
    """
    out: Dict[str, str] = {}
    wanted = {s.lower() for s in subscription_ids if s}

    try:
        sc = SubscriptionClient(get_cred())
        for sub in sc.subscriptions.list():
            sid = (getattr(sub, "subscription_id", "") or "").lower()
            if sid and sid in wanted:
                out[sid] = getattr(sub, "display_name", "") or sid
    except Exception as ex:
        print(f"⚠ 구독명 조회 실패(SubscriptionClient): {ex}")

    # fallback
    for s in subscription_ids:
        if s and s.lower() not in out:
            out[s.lower()] = s

    return out


def build_summary_rows(
    all_rows_by_category: List[List[Dict[str, Any]]],
    sub_name_map: Dict[str, str]
) -> List[Dict[str, Any]]:
    """
    모든 시트 rows를 받아 Summary 시트를 만들고,
    구독명/리소스그룹/리소스타입/리소스명 + (리소스별 IP) 를 정리.

    ✅ Summary의 IP 컬럼 규칙
    - VM: NIC Private IP  (row["Private IP"])
    - AppGW: Public/Private Frontend IP (row["Public IP"], row["Private IP"]) -> "Public:x / Private:y"
    - Load Balancer: Frontend IP (row["Frontend IP"])
    - Private Endpoint: Private IP (row["Private IP"])
    - 그 외: 공란
    """

    def _pick_ip_for_summary(row: Dict[str, Any], rtype: str) -> str:
        """
        Summary의 IP/FQDN 컬럼에 들어갈 값 선택 규칙

        우선순위:
        1) Public/Private/Frontend IP
        2) FQDN / URL / Endpoint / Host
        """

        # =========================
        # 1. IP 계열 (최우선)
        # =========================
        pub = (row.get("Public IP") or "").strip()
        priv = (row.get("Private IP") or "").strip()
        feip = (row.get("Frontend IP") or "").strip()

        if pub or priv:
            parts = []
            if pub:
                parts.append(f"Public:{pub}")
            if priv:
                parts.append(f"Private:{priv}")
            return " / ".join(parts)

        if feip:
            return feip

        # =========================
        # 2. FQDN / URL 계열
        # =========================
        for key in [
            # ACR
            "Login Server",

            # ACA
            "Application URL",

            # Function App / App Service
            "Default Domain",

            # AI Search
            "URL",

            # Document Intelligence / PostgreSQL / MySQL / Redis
            "Endpoint",

            # Cosmos DB
            "URI",

            # SQL MI
            "Host",

            # Azure OpenAI
            "Endpoint",

            # Key Vault
            "Vault URI",
        ]:
            v = (row.get(key) or "").strip()
            if v:
                return v

        return ""

    out: List[Dict[str, Any]] = []
    seen: Set[Tuple[str, str, str, str, str]] = set()

    for rows in all_rows_by_category:
        for r in rows:
            sub_id = (r.get("SubscriptionID") or "").strip()
            rg = (r.get("ResourceGroup") or "").strip()
            rtype = (r.get("Type") or r.get("ResourceType") or "").strip()
            name = (r.get("ResourceName") or "").strip()

            sub_name = sub_name_map.get(sub_id.lower(), sub_id)

            ip_val = _pick_ip_for_summary(r, rtype)

            key = (sub_name, rg, rtype, name, ip_val)
            if key in seen:
                continue
            seen.add(key)

            out.append({
                "Subscription Name": sub_name,
                "ResourceGroup": rg,
                "ResourceType": rtype,
                "ResourceName": name,
                "IP/FQDN": ip_val,   
            })

    out.sort(key=lambda x: (x["Subscription Name"], x["ResourceGroup"], x["ResourceType"], x["ResourceName"]))
    return out




# =========================
# 3. ARG Queries / API Queries
# =========================

# [VM] Azure Virtual Machines (ARG)
def query_azure_vms_basic(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.compute/virtualmachines'
| extend
    vmSize   = tostring(properties.hardwareProfile.vmSize),
    osType   = tostring(properties.storageProfile.osDisk.osType),
    osOffer  = tostring(properties.storageProfile.imageReference.offer),
    osSku    = tostring(properties.storageProfile.imageReference.sku),
    osDiskId = tolower(tostring(properties.storageProfile.osDisk.managedDisk.id)),
    nicId    = tolower(tostring(properties.networkProfile.networkInterfaces[0].id))
| extend dataDisks = properties.storageProfile.dataDisks
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    vmSize,
    osType,
    osOffer,
    osSku,
    osDiskId,
    nicId,
    dataDisks
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Virtual Machine 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

def query_azure_vmss_basic(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.compute/virtualmachinescalesets'
| extend
    skuName = tostring(sku.name),
    instanceCount = toint(sku.capacity),
    orchestrationMode = tostring(properties.orchestrationMode),
    osType = tostring(properties.virtualMachineProfile.storageProfile.osDisk.osType),
    osOffer = tostring(properties.virtualMachineProfile.storageProfile.imageReference.offer),
    osSku = tostring(properties.virtualMachineProfile.storageProfile.imageReference.sku),
    osDiskSizeGB = toint(properties.virtualMachineProfile.storageProfile.osDisk.diskSizeGB),
    osDiskStorageType = tostring(properties.virtualMachineProfile.storageProfile.osDisk.managedDisk.storageAccountType),
    dataDisks = properties.virtualMachineProfile.storageProfile.dataDisks
| project
    subscriptionId, resourceGroup, name, type, location,
    skuName, instanceCount, orchestrationMode,
    osType, osOffer, osSku,
    osDiskSizeGB, osDiskStorageType,
    dataDisks
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 VMSS 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]


# [VM] NIC Private IP (ARG)
def query_nic_private_ips(subscription_ids: List[str], rg_name: str) -> Dict[str, str]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.network/networkinterfaces'
| extend nicId = tolower(id)
| mv-expand ipconfig = properties.ipConfigurations
| extend privateIP = tostring(ipconfig.properties.privateIPAddress)
| summarize privateIP = take_anyif(privateIP, isnotempty(privateIP)) by nicId
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name)) 
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 NIC(private IP) 조회 중...")
    result = client.resources(req)
    return {str(r["nicId"]): str(r.get("privateIP","")) for r in result.data}


# [VM] Managed Disks (ARG)
def query_disks(subscription_ids: List[str], rg_name: str) -> Dict[str, Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.compute/disks'
| project
    diskId = tolower(id),
    diskName = name,
    diskSku = tostring(sku.name),
    diskSizeGiB = toint(properties.diskSizeGB)
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))  # ✅ 추가
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Disk 조회 중...")
    result = client.resources(req)
    out: Dict[str, Dict[str, Any]] = {}
    for r in result.data:
        out[str(r["diskId"])] = {
            "name": r.get("diskName", ""),
            "sku": r.get("diskSku", ""),
            "size": r.get("diskSizeGiB", "")
        }
    return out

# [AppGW] Application Gateways (ARG)
def query_azure_appgws(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.network/applicationgateways'
| extend
    skuNameRaw  = coalesce(tostring(sku.name), tostring(properties.sku.name)),
    skuTierRaw  = coalesce(tostring(sku.tier), tostring(properties.sku.tier)),
    minCapacity = toint(properties.autoscaleConfiguration.minCapacity),
    maxCapacity = toint(properties.autoscaleConfiguration.maxCapacity)
| mv-expand fe = properties.frontendIPConfigurations
| extend
    publicIPId = tolower(tostring(fe.properties.publicIPAddress.id)),
    privateIP  = tostring(fe.properties.privateIPAddress)
| summarize
    publicIPId = take_anyif(publicIPId, isnotempty(publicIPId)),
    privateIP  = take_anyif(privateIP,  isnotempty(privateIP)),
    skuNameRaw = any(skuNameRaw),
    skuTierRaw = any(skuTierRaw),
    minCapacity = any(minCapacity),
    maxCapacity = any(maxCapacity)
  by subscriptionId, resourceGroup, name, type, location
| extend
    skuType = case(
        (skuNameRaw has 'WAF' and skuNameRaw has 'v2') or (skuTierRaw has 'WAF' and skuTierRaw has 'v2'), 'WAFv2',
        (skuNameRaw has 'Standard' and skuNameRaw has 'v2') or (skuTierRaw has 'Standard' and skuTierRaw has 'v2'), 'StandardV2',
        isnotempty(skuNameRaw), skuNameRaw,
        skuTierRaw
    )
| join kind=leftouter (
    Resources
    | where type =~ 'microsoft.network/publicipaddresses'
    | extend publicIPId = tolower(id)
    | project publicIPId, publicIPAddress = tostring(properties.ipAddress)
) on publicIPId
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    skuType,
    minCapacity,
    maxCapacity,
    publicIPAddress,
    privateIP
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Application Gateway 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [ACR] Container Registry (ARG)
def query_azure_acrs(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.containerregistry/registries'
| extend skuName = tostring(sku.name)
| extend loginServer = tostring(properties.loginServer)
| project subscriptionId, resourceGroup, name, type, location, skuName, loginServer
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Container Registry 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [PostgreSQL] Azure Database for PostgreSQL (ARG)
def query_azure_postgresql(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.dbforpostgresql/flexibleservers'
   or type =~ 'microsoft.dbforpostgresql/servers'
| extend
    skuName = tostring(sku.name),
    version = tostring(properties.version),
    storageSizeGB_flex = toint(properties.storage.storageSizeGB),
    storageIops_flex   = toint(properties.storage.iops),
    haMode_flex        = tostring(properties.highAvailability.mode),
    haState_flex       = tostring(properties.highAvailability.state),
    storageMB_single   = toint(properties.storageProfile.storageMB)
| extend
    storageSizeGB = iif(isnotnull(storageSizeGB_flex) and storageSizeGB_flex > 0,
                        storageSizeGB_flex,
                        iif(isnotnull(storageMB_single) and storageMB_single > 0, toint(storageMB_single / 1024), int(null))),
    storageIops   = storageIops_flex,
    HA = case(
        isnotempty(haMode_flex) and isnotempty(haState_flex), strcat(haMode_flex, ' (', haState_flex, ')'),
        isnotempty(haMode_flex), haMode_flex,
        'N/A'
    )
| extend fqdn = tostring(properties.fullyQualifiedDomainName)
| project subscriptionId, resourceGroup, name, type, location, skuName, version, storageSizeGB, storageIops, HA, fqdn
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure Database for PostgreSQL Flexible 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [KeyVault] Key Vaults (ARG)
def query_azure_keyvaults(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.keyvault/vaults'
| extend
    skuName = tostring(properties.sku.name),
    skuFamily = tostring(properties.sku.family),
    vaultUri = tostring(properties.vaultUri)
| extend
    sku = case(
        isnotempty(skuName) and isnotempty(skuFamily), strcat(skuName, " (", skuFamily, ")"),
        isnotempty(skuName), skuName,
        isnotempty(skuFamily), skuFamily,
        "N/A"
    )
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    sku,
    vaultUri
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Key Vault 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Container Apps] Container Apps + Workload Profile (ARG)
def query_azure_containerapps_with_workload_profiles(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()

    rg = (rg_name or "").strip()
    rg_esc = esc_kql_str(rg)

    # RG 필터를 쿼리 안에서 명시적으로 적용 (add_rg_filter 사용 X)
    query = normalize_arg_query(f"""
Resources
| where resourceGroup =~ '{rg_esc}'
| where type =~ 'microsoft.app/containerapps'
| extend
    appLocation = tostring(location),
    envId = tolower(tostring(properties.managedEnvironmentId)),
    workloadProfileName = tostring(properties.workloadProfileName),
    appFqdn = coalesce(
        tostring(properties.configuration.ingress.fqdn),
        tostring(properties.latestRevisionFqdn),
        tostring(properties.ingress.fqdn)
    )
| join kind=leftouter (
    Resources
    | where resourceGroup =~ '{rg_esc}'
    | where type =~ 'microsoft.app/managedenvironments'
    | extend envId = tolower(id)
    | mv-expand wp = properties.workloadProfiles
    | project
        envId,
        envName = name,
        wpName = tostring(wp.name),
        wpType = tostring(wp.workloadProfileType),
        envLocation = tostring(location)
) on envId
| where isempty(workloadProfileName) or workloadProfileName == wpName
| summarize
    workloadProfileName = any(workloadProfileName),
    workloadProfileType = take_anyif(wpType, isnotempty(wpType)),
    envLocation = take_anyif(envLocation, isnotempty(envLocation)),
    envName = take_anyif(envName, isnotempty(envName)),  
    appLocation = any(appLocation),
    type = any(type),
    appFqdn = take_anyif(appFqdn, isnotempty(appFqdn))
  by subscriptionId, resourceGroup, name
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location = appLocation,
    envName,                      
    workloadProfileName,
    workloadProfileType,
    envLocation,
    appFqdn
""")

    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Container Apps 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]



# [Container Apps] Workload Profile Types (ARM REST) - location별 CPU/Memory 매핑
def fetch_available_workload_profile_types(subscription_id: str, location: str, rg_name: str) -> Dict[str, Dict[str, Any]]:
    api_version = "2025-07-01"
    url = (
        f"https://management.azure.com/subscriptions/{subscription_id}"
        f"/providers/Microsoft.App/locations/{location}"
        f"/availableManagedEnvironmentsWorkloadProfileTypes"
        f"?api-version={api_version}"
    )
    headers = {"Authorization": f"Bearer {get_arm_token()}"}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()

    out: Dict[str, Dict[str, Any]] = {}
    for item in data.get("value", []):
        name = item.get("name", "")
        props = item.get("properties", {}) or {}
        out[name] = {
            "cores": props.get("cores"),
            "memoryGiB": props.get("memoryGiB"),
        }
    return out

# [AKS] Client (SDK)
def get_aks_client(subscription_id: str) -> ContainerServiceClient:
    return ContainerServiceClient(get_cred(), subscription_id)

# [AKS] NodePools (SDK)
def query_aks_nodepools_api(subscription_id: str, rg_name: str) -> List[Dict[str, Any]]:
    print("👉 AKS(Nodepool) 조회 중...")
    client = get_aks_client(subscription_id)

    out: List[Dict[str, Any]] = []
    rg_target = (rg_name or "").strip().lower()

    for mc in client.managed_clusters.list():
        rg = ""
        if getattr(mc, "id", None):
            parts = mc.id.split("/")
            for idx, p in enumerate(parts):
                if p.lower() == "resourcegroups" and idx + 1 < len(parts):
                    rg = parts[idx + 1]
                    break
        if rg_target and rg.lower() != rg_target:  
            continue

        cluster_name = mc.name or ""
        cluster_type = "microsoft.containerservice/managedclusters"
        location = getattr(mc, "location", None) or ""
        cluster_os_hint = "Linux" if getattr(mc, "linux_profile", None) else ""

        for ap in client.agent_pools.list(rg, cluster_name):
            os_type = getattr(ap, "os_type", None) or ""
            os_sku = getattr(ap, "os_sku", None) or ""
            os_str = " / ".join([x for x in [os_type, os_sku] if x]) or cluster_os_hint

            node_count = getattr(ap, "count", None)
            enable_autoscaling = getattr(ap, "enable_auto_scaling", None)

            if (node_count is None or node_count == 0) and enable_autoscaling:
                min_count = getattr(ap, "min_count", None)
                if isinstance(min_count, int) and min_count > 0:
                    node_count = min_count

            vm_size = getattr(ap, "vm_size", None) or ""
            os_disk_size_gb = getattr(ap, "os_disk_size_gb", None)

            out.append({
                "subscriptionId": subscription_id,
                "resourceGroup": rg,
                "clusterName": cluster_name,
                "type": cluster_type,
                "location": location,
                "os": os_str,
                "nodepoolName": ap.name or "",
                "nodeCount": node_count if node_count is not None else "",
                "vmSize": vm_size,
                "osDiskSizeGB": os_disk_size_gb if os_disk_size_gb is not None else "",
            })

    return out

# [Storage] Storage Accounts (ARG)
def query_storage_accounts_blob_config(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.storage/storageaccounts'
| project subscriptionId, resourceGroup, name, type, location, id, sku, kind, properties
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Storage Account 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Azure Files] File Shares (ARM REST)
def list_file_shares(subscription_id: str, resource_group: str, account_name: str) -> List[Dict[str, Any]]:
    token = get_arm_token()
    headers = {"Authorization": f"Bearer {token}"}
    api_version = "2023-01-01"

    url = (
        f"https://management.azure.com/subscriptions/{subscription_id}"
        f"/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Storage/storageAccounts/{account_name}"
        f"/fileServices/default/shares"
    )
    r = requests.get(url, headers=headers, params={"api-version": api_version}, timeout=30)
    r.raise_for_status()
    return r.json().get("value", [])

def get_file_share_stats(subscription_id: str, resource_group: str, account_name: str, share_name: str) -> int:
    """
    Azure Files Share 사용량 조회 (bytes)
    - 성공 시 shareUsageBytes(int) 반환
    - 실패 시 -1 반환
    """
    token = get_arm_token()
    headers = {"Authorization": f"Bearer {token}"}
    api_version = "2023-01-01"

    url = (
        f"https://management.azure.com/subscriptions/{subscription_id}"
        f"/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Storage/storageAccounts/{account_name}"
        f"/fileServices/default/shares/{share_name}/stats"
    )
    try:
        r = requests.get(url, headers=headers, params={"api-version": api_version}, timeout=30)
        r.raise_for_status()
        data = r.json() or {}
        v = data.get("shareUsageBytes", None)
        return int(v) if v is not None else -1
    except Exception:
        return -1


def bytes_to_gb(b: int) -> str:
    if b is None or b < 0:
        return ""
    return str(round(b / (1024 ** 3), 2))


# [AI Search] Search Services (ARG)
def query_azure_ai_search(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.search/searchservices'
| extend
    skuName = tostring(sku.name),
    replicaCount = toint(properties.replicaCount),
    partitionCount = toint(properties.partitionCount),
    hostName = tostring(properties.hostName)
| extend url = iif(isnotempty(hostName), strcat('https://', hostName), strcat('https://', name, '.search.windows.net'))
| project subscriptionId, resourceGroup, name, type, location, skuName, replicaCount, partitionCount, url
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 AI Search 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Document Intelligence] CognitiveServices Accounts (ARG)
def query_azure_document_intelligence(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.cognitiveservices/accounts'
| extend kindName = tostring(kind)
| where kindName =~ 'FormRecognizer' or kindName =~ 'DocumentIntelligence'
| extend skuName = tostring(sku.name)
| extend endpoint = tostring(properties.endpoint)
| project subscriptionId, resourceGroup, name, type, location, kindName, skuName, endpoint
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Document Intelligence 조회 중...")

    try:
        result = client.resources(req)
        return [dict(row) for row in result.data]
    except Exception as ex:
        print(f"⚠ Document Intelligence 조회 실패: {ex}")
        return []

# [AML] Machine Learning Workspaces (ARG)
def query_azure_machine_learning_workspaces(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.machinelearningservices/workspaces'
| extend
    acrId = tostring(properties.containerRegistry),
    kvId  = tostring(properties.keyVault),
    aiId  = tostring(properties.applicationInsights)
| extend
    acrName = extract(@'/registries/([^/]+)$', 1, acrId),
    kvName  = extract(@'/vaults/([^/]+)$',     1, kvId),
    aiName  = extract(@'/components/([^/]+)$', 1, aiId)
| project subscriptionId, resourceGroup, name, type, location, id, acrName, kvName, aiName
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure ML(Workspace) 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Event Hubs] EventHub Namespaces (ARG)
def query_azure_eventhub_namespaces(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.eventhub/namespaces'
| extend
    tier = coalesce(tostring(sku.name), tostring(sku.tier)),
    skuCapacity = toint(sku.capacity),
    throughputUnits = toint(properties.throughputUnits),
    autoInflate = tobool(properties.isAutoInflateEnabled),
    maxThroughputUnits = toint(properties.maximumThroughputUnits),
    premiumPartitions = toint(properties.premiumMessagingPartitions)
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    tier,
    throughputUnits,
    autoInflate,
    maxThroughputUnits,
    premiumPartitions,
    skuCapacity
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Event Hubs Namespace 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Azure OpenAI] CognitiveServices Accounts 중 OpenAI (ARG)
def query_azure_openai(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.cognitiveservices/accounts'
| extend kindName = tostring(kind)
| where kindName =~ 'OpenAI'
| extend
    skuName = tostring(sku.name),
    endpoint = tostring(properties.endpoint)
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    skuName,
    endpoint
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure OpenAI 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [MySQL] Azure Database for MySQL Flexible Server (ARG)
def query_azure_mysql_flexible(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.dbformysql/flexibleservers'
| extend
    skuName = tostring(sku.name),
    version = tostring(properties.version),
    storageSizeGB = toint(properties.storage.storageSizeGB),
    storageIops   = toint(properties.storage.iops),
    haMode        = tostring(properties.highAvailability.mode),
    haState       = tostring(properties.highAvailability.state)
| extend
    HA = case(
        isnotempty(haMode) and isnotempty(haState), strcat(haMode, ' (', haState, ')'),
        isnotempty(haMode), haMode,
        'N/A'
    )
| extend fqdn = tostring(properties.fullyQualifiedDomainName)
| project subscriptionId, resourceGroup, name, type, location, skuName, version, storageSizeGB, storageIops, HA, fqdn
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure Database for MySQL Flexible 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Azure Managed Redis] (ARG) - Microsoft.Cache/redisEnterprise
def query_azure_managed_redis(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.cache/redisenterprise'
| extend
    skuName = tostring(sku.name),
    ha = tostring(properties.highAvailability)
| extend
    tier = tostring(split(skuName, "_")[0]),
    instance = tostring(split(skuName, "_")[1])
| extend host = tostring(properties.hostName)
| project subscriptionId, resourceGroup, name, type, location, tier, instance, ha, skuName, host
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure Managed Redis 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Azure DocumentDB (MongoDB vCore)] (ARG) - 목록만 (id 확보)
def query_azure_documentdb_list(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.documentdb/mongoclusters'
| project subscriptionId, resourceGroup, name, type, location, id
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure DocumentDB (with MongoDB compatibility) 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Azure DocumentDB (MongoDB vCore)] (ARM REST) - 상세 properties 조회
def fetch_documentdb_details(subscription_id: str, resource_group: str, cluster_name: str) -> Dict[str, Any]:
    api_version = "2025-09-01"
    url = (
        f"https://management.azure.com/subscriptions/{subscription_id}"
        f"/resourceGroups/{resource_group}"
        f"/providers/Microsoft.DocumentDB/mongoClusters/{cluster_name}"
        f"?api-version={api_version}"
    )

    headers = {"Authorization": f"Bearer {get_arm_token()}"}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()

    data = r.json() or {}
    props = data.get("properties") or {}

    compute = props.get("compute") or {}
    sharding = props.get("sharding") or {}
    storage = props.get("storage") or {}
    ha = props.get("highAvailability") or {}

    return {
        "clusterTier": (compute.get("tier") or ""),
        "shardCount": sharding.get("shardCount", ""),
        "storageGb": storage.get("sizeGb", ""),
        "haMode": (ha.get("targetMode") or ha.get("mode") or ""),
    }

# [Azure CosmosDB (Core/SQL/MongoDB API(RU))] (ARG)
def query_cosmosdb_mongo_accounts(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.documentdb/databaseaccounts'
| extend kindName = tostring(kind)
| extend capabilities = properties.capabilities
| extend isMongo = iif(
    kindName has 'Mongo',
    true,
    iif(tostring(capabilities) has 'EnableMongo', true, false)
)
| where isMongo == true
| extend
    enableMultiWrite = tobool(properties.enableMultipleWriteLocations),
    locs = properties.locations,
    totalThroughputLimit = toint(properties['capacity']['totalThroughputLimit'])
| extend uri = tostring(properties.documentEndpoint)
| project subscriptionId, resourceGroup, name, type, location, enableMultiWrite, locs, totalThroughputLimit, uri
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure CosmosDB for MongoDB account (RU) 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Azure CosmosDB for PostgreSQL] (ARG) - serverGroupsv2
def query_cosmosdb_postgresql(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.dbforpostgresql/servergroupsv2'
| extend
    nodeCount = toint(properties['nodeCount']),
    coordEdition = tostring(properties['coordinatorServerEdition']),
    coordVCores  = toint(properties['coordinatorVCores']),
    nodeEdition  = tostring(properties['nodeServerEdition']),
    nodeVCores   = toint(properties['nodeVCores'])
| project
    subscriptionId, resourceGroup, name, type, location,
    nodeCount, coordEdition, coordVCores, nodeEdition, nodeVCores
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure CosmosDB for PostgreSQL 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [SQL Managed Instance] (ARG)
def query_azure_sql_managed_instances(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.sql/managedinstances'
| extend
    skuName = tostring(sku.name),
    vcores  = toint(properties.vCores),
    storageGB = toint(properties.storageSizeInGB),
    zoneRedundant = tobool(properties.zoneRedundant)
| extend host = tostring(properties.fullyQualifiedDomainName)
| project subscriptionId, resourceGroup, name, type, location, skuName, vcores, storageGB, zoneRedundant, host
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Azure SQL Managed Instance 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Virtual Network] (ARG) - vNet/subnet
def query_virtual_networks_with_subnets(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.network/virtualnetworks'
| extend
    vnetName = name,
    vnetPrefixes = properties.addressSpace.addressPrefixes
| mv-expand subnet = properties.subnets
| extend
    subnetName = tostring(subnet.name),
    subnetPrefix = coalesce(
        tostring(subnet.properties.addressPrefix),
        tostring(subnet.properties.addressPrefixes[0])
    ),
    nsgId = tostring(subnet.properties.networkSecurityGroup.id),
    udrId = tostring(subnet.properties.routeTable.id)
| extend
    nsgName = extract(@'/networkSecurityGroups/([^/]+)$', 1, nsgId),
    udrName = extract(@'/routeTables/([^/]+)$', 1, udrId)
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    vnetName,
    vnetPrefixes,
    subnetName,
    subnetPrefix,
    nsgName,
    udrName
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Virtual Network(vNet+Subnet+NSG+UDR) 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Private Endpoint] (ARG)
def query_private_endpoints(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.network/privateendpoints'
| extend
    plsConn    = properties.privateLinkServiceConnections,
    manualConn = properties.manualPrivateLinkServiceConnections
| extend
    plsId1 = tostring(plsConn[0].properties.privateLinkServiceId),
    plsId2 = tostring(manualConn[0].properties.privateLinkServiceId)
| extend targetResourceId = coalesce(plsId1, plsId2)
| extend targetResourceName = extract(@'/([^/]+)$', 1, targetResourceId)
| extend connected = case(
    isnotempty(targetResourceName) and isnotempty(targetResourceId), strcat(targetResourceName, ' (', targetResourceId, ')'),
    isnotempty(targetResourceId), targetResourceId,
    'N/A'
)
| extend
    peId = tolower(id),
    // ⭐ PE가 붙는 NIC id 확보 (가장 안정적)
    nicId = tolower(tostring(properties.networkInterfaces[0].id))
| join kind=leftouter (
    Resources
    | where type =~ 'microsoft.network/networkinterfaces'
    | extend nicId = tolower(id)
    | mv-expand ipconf = properties.ipConfigurations
    | extend privateIP = tostring(ipconf.properties.privateIPAddress)
    | summarize privateIP = take_anyif(privateIP, isnotempty(privateIP)) by nicId
) on nicId
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    privateIP,
    connected
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Private Endpoint 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]


# [Log Analytics] (ARG)
def query_log_analytics_workspaces(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.operationalinsights/workspaces'
| extend
    skuName = tostring(properties.sku.name),
    retentionDays = toint(properties.retentionInDays),
    publicNetworkAccess = tostring(properties.publicNetworkAccessForIngestion)
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    skuName,
    retentionDays,
    publicNetworkAccess
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Log Analytics Workspace 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Application Insights] (ARG)
def query_application_insights(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.insights/components'
| extend
    appType = tostring(properties.Application_Type),
    workspaceResourceId = tolower(tostring(properties.WorkspaceResourceId))
| project
    subscriptionId,
    resourceGroup,
    name,
    type,
    location,
    appType,
    workspaceResourceId
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Application Insights 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Function App] (ARG)
def query_function_apps(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.web/sites'
| extend kindStr = tolower(tostring(kind))
| where kindStr has 'functionapp'
| where not(kindStr has 'workflowapp')   // ✅ Logic App Standard 제외
| extend
    defaultHost = tostring(properties.defaultHostName),
    reserved    = tobool(properties.reserved),  // true면 Linux
    planId      = tolower(tostring(properties.serverFarmId)),
    fxVersion   = tostring(properties.siteConfig.linuxFxVersion),
    windowsFx   = tostring(properties.siteConfig.windowsFxVersion),
    funcVer     = tostring(properties.siteConfig.functionsExtensionVersion)
| extend
    os = iif(reserved == true, "Linux", "Windows"),
    runtime = case(
        isnotempty(funcVer), strcat("Functions ", funcVer),
        isnotempty(fxVersion), fxVersion,
        isnotempty(windowsFx), windowsFx,
        "N/A"
    )
| project
    subscriptionId, resourceGroup, name, type, location,
    defaultHost, os, planId, runtime
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Function Apps 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [Logic App] (ARG)
def query_logic_apps(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()

    rg = (rg_name or "").strip()
    rg_filter = ""
    if rg:
        rg_filter = f"| where resourceGroup =~ '{esc_kql_str(rg)}'"

    query = normalize_arg_query(f"""
(
    Resources
    {rg_filter}
    | where type =~ 'microsoft.web/sites'
    | extend kindStr = tolower(tostring(kind))
    | where kindStr has 'workflowapp'
    | extend
        defaultHost = tostring(properties.defaultHostName),
        reserved    = tobool(properties.reserved),
        planId      = tolower(tostring(properties.serverFarmId)),
        fxVersion   = tostring(properties.siteConfig.linuxFxVersion),
        windowsFx   = tostring(properties.siteConfig.windowsFxVersion),
        funcVer     = tostring(properties.siteConfig.functionsExtensionVersion)
    | extend
        os = iif(reserved == true, 'Linux', 'Windows'),
        runtime = case(
            isnotempty(funcVer), strcat('Functions ', funcVer),
            isnotempty(fxVersion), fxVersion,
            isnotempty(windowsFx), windowsFx,
            'N/A'
        )
    | project
        subscriptionId, resourceGroup, name, type, location,
        defaultHost, os, planId, runtime
)
| union
(
    Resources
    {rg_filter}
    | where type =~ 'microsoft.logic/workflows'
    | extend accessEndpoint = tostring(properties.accessEndpoint)
    | extend
        defaultHost = accessEndpoint,
        os = 'N/A',
        planId = '',
        runtime = 'Consumption'
    | project
        subscriptionId, resourceGroup, name, type, location,
        defaultHost, os, planId, runtime
)
""")

    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 Logic Apps 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]

# [App Service Plan] (ARG)
def query_app_service_plans(subscription_ids: List[str], rg_name: str) -> Dict[str, Dict[str, str]]:
    """
    App Service Plan(Server Farm) id -> { name, pricing_tier }
    pricing_tier = sku.tier + sku.name (예: PremiumV3 P1v3)
    """
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.web/serverfarms'
| extend planId = tolower(id)
| extend skuName = tostring(sku.name), skuTier = tostring(sku.tier)
| extend pricingTier = case(
    isnotempty(skuTier) and isnotempty(skuName), strcat(skuTier, " ", skuName),
    isnotempty(skuName), skuName,
    isnotempty(skuTier), skuTier,
    "N/A"
)
| project planId, planName=name, pricingTier
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 App Service Plan 조회 중...")
    result = client.resources(req)

    out: Dict[str, Dict[str, str]] = {}
    for r in result.data:
        pid = str(r.get("planId") or "").lower()
        if not pid:
            continue
        out[pid] = {
            "name": str(r.get("planName") or ""),
            "pricingTier": str(r.get("pricingTier") or "N/A"),
        }
    return out

# [Load Balancer] (ARM)
def query_load_balancers_rest(subscription_id: str, rg_name: str) -> List[Dict[str, Any]]:
    print(f"👉 Azure Load Balancer 조회 중...")

    api_version = "2023-09-01"
    url = (
        f"https://management.azure.com/subscriptions/{subscription_id}"
        f"/resourceGroups/{rg_name}"
        f"/providers/Microsoft.Network/loadBalancers"
    )

    data = arm_get(url, api_version=api_version)
    lbs = data.get("value", []) or []

    pip_cache: Dict[str, str] = {}
    out: List[Dict[str, Any]] = []

    for lb in lbs:
        lb_name = lb.get("name", "")
        lb_type = lb.get("type", "")
        lb_loc  = lb.get("location", "")

        sku = lb.get("sku") or {}
        sku_name = sku.get("name", "") if isinstance(sku, dict) else ""
        sku_tier = sku.get("tier", "") if isinstance(sku, dict) else ""

        props = lb.get("properties") or {}

        # Frontend
        fe_names: List[str] = []
        fe_ips: List[str] = []
        for fe in (props.get("frontendIPConfigurations") or []):
            fe_names.append(fe.get("name", "") or "")
            fe_props = fe.get("properties") or {}
            priv_ip = fe_props.get("privateIPAddress", "") or ""
            pip_id = ((fe_props.get("publicIPAddress") or {}) or {}).get("id", "") or ""
            pub_ip = get_public_ip_by_id(pip_id, pip_cache) if pip_id else ""
            fe_ips.append(pub_ip or priv_ip or "")

        # Backend pools
        bp_names = [bp.get("name", "") or "" for bp in (props.get("backendAddressPools") or [])]

        # Rules
        rule_names = [r.get("name", "") or "" for r in (props.get("loadBalancingRules") or [])]

        # Probes
        probe_names = [p.get("name", "") or "" for p in (props.get("probes") or [])]

        out.append({
            "subscriptionId": subscription_id,
            "resourceGroup": rg_name,
            "name": lb_name,
            "type": lb_type,
            "location": lb_loc,
            "SKU": sku_name,
            "Tier": sku_tier,
            "FrontendNames": [x for x in fe_names if x],
            "FrontendIPs": [x for x in fe_ips if x],
            "BackendPoolNames": [x for x in bp_names if x],
            "LBRuleNames": [x for x in rule_names if x],
            "ProbeNames": [x for x in probe_names if x],
        })

    return out

def query_user_assigned_managed_identities(subscription_ids: List[str], rg_name: str) -> List[Dict[str, Any]]:
    client = get_arg_client()
    base = normalize_arg_query(r"""
Resources
| where type =~ 'microsoft.managedidentity/userassignedidentities'
| project subscriptionId, resourceGroup, name, type, location
""")
    query = normalize_arg_query(add_rg_filter(base, rg_name))
    req = QueryRequest(subscriptions=subscription_ids, query=query)
    print("👉 User Assigned Managed Identity 조회 중...")
    result = client.resources(req)
    return [dict(row) for row in result.data]


# =========================
# 4. Helper Parsers
# =========================

def build_vm_size_map(subscription_id: str, locations: List[str]) -> Dict[str, Dict[str, Any]]:
    compute_client = ComputeManagementClient(get_cred(), subscription_id)
    size_map: Dict[str, Dict[str, Any]] = {}

    for loc in {l for l in locations if l}:
        try:
            print(f"👉 {subscription_id} / {loc} 에서 VM SKU 정보 조회 중...")
            for s in compute_client.virtual_machine_sizes.list(loc):
                if s.name not in size_map:
                    size_map[s.name] = {
                        "vCPU": s.number_of_cores,
                        "Memory": round(s.memory_in_mb / 1024, 1),
                    }
        except Exception as ex:
            print(f"⚠ VM SKU 조회 실패 (location={loc}): {ex}")

    return size_map


def parse_redundancy(sku_name: str) -> str:
    if not sku_name:
        return ""
    parts = sku_name.split("_")
    return parts[-1] if len(parts) >= 2 else sku_name

def parse_namespace(hns_value: str) -> str:
    if str(hns_value).lower() == "true":
        return "계층 구조 네임스페이스"
    if str(hns_value).lower() == "false":
        return "단일 계층 네임스페이스"
    return ""

def parse_redundancy_from_sku_name(sku_name: str) -> str:
    if not sku_name:
        return ""
    parts = sku_name.split("_")
    return parts[-1] if len(parts) >= 2 else sku_name

def parse_media_tier(sku_tier: str, kind: str, sku_name: str) -> str:
    t = (sku_tier or "").lower()
    k = (kind or "").lower()
    sn = (sku_name or "").lower()

    if "premium" in t or k == "filestorage" or sn.startswith("premium"):
        return "SSD"
    if "standard" in t or sn.startswith("standard"):
        return "HDD"
    return ""

def map_access_tier(tier: str) -> str:
    t = (tier or "").lower()
    if t == "transactionoptimized":
        return "트랜잭션 최적화됨"
    if t == "hot":
        return "핫"
    if t == "cool":
        return "쿨"
    return tier or ""

def parse_billing_model(share_props: Dict[str, Any], sku_tier: str, kind: str) -> str:
    """
    Azure Files Billing Model:
    - Standard(표준): 종량제 / 프로비전v2
    - Premium(프리미엄, FileStorage): 프로비전v1 / 프로비전v2

    share_props에 버전 정보가 없으면 Premium에서도 v1/v2 확정이 불가할 수 있어
    -> '프로비전(버전미확인)'로 표시
    """

    # 0) 입력 정리
    props = share_props if isinstance(share_props, dict) else {}
    sku_tier_l = (sku_tier or "").strip().lower()
    kind_l = (kind or "").strip().lower()

    is_premium = (sku_tier_l == "premium") or (kind_l == "filestorage")  # FileStorage는 Premium 계열로 간주

    # 1) 버전/모델 힌트 후보 키들 (환경/버전에 따라 다르게 내려오는 경우 대비)
    candidates = [
        props.get("provisionedBillingModelVersion", None),
        props.get("provisionedBillingModel", None),
        props.get("billingModelVersion", None),
        props.get("billingModel", None),
    ]
    raw = ""
    for c in candidates:
        if c not in (None, "", "null"):
            raw = str(c).strip()
            break

    raw_l = raw.lower()

    # 2) Premium: 프로비전v1 / 프로비전v2
    if is_premium:
        if raw:
            # v2 우선
            if ("v2" in raw_l) or ("version2" in raw_l) or (raw_l == "2") or (raw_l.endswith("2")):
                return "프로비전v2"
            if ("v1" in raw_l) or ("version1" in raw_l) or (raw_l == "1") or (raw_l.endswith("1")):
                return "프로비전v1"

            # 값은 있는데 v1/v2 판별이 안 되면
            return f"프로비전({raw})"

        # Premium인데 버전 정보가 안 내려옴 -> v1/v2 확정 불가
        return "프로비전(버전미확인)"

    # 3) Standard: 종량제 / 프로비전v2
    # Standard에서 provisioned가 잡히면 v2로 취급 (너가 말한 조건 반영)
    if raw:
        # 표준에서 프로비전은 v2만 있다고 가정
        return "프로비전v2"

    return "종량제"


def parse_sqlmi_service_tier(sku_name: str) -> str:
    s = (sku_name or "").upper()
    if s.startswith("GP_"):
        return "범용"
    if s.startswith("BC_"):
        return "중요 비즈니스용"
    return ""

def parse_sqlmi_hw_generation(sku_name: str) -> str:
    s = (sku_name or "").upper()
    if "GEN5" in s:
        return "표준"
    if "G8IM" in s:
        return "프리미엄"
    if "G8IH" in s:
        return "프리미엄 메모리 최적화"
    return ""

def estimate_sqlmi_memory_gb(hw_gen: str, vcores: Any) -> str:
    try:
        vc = int(vcores)
        if vc <= 0:
            return ""
    except Exception:
        return ""

    hw = (hw_gen or "").strip()

    if hw == "표준":
        mem = 5.1 * vc
        mem = min(mem, 408.0)
        return str(round(mem, 1))

    if hw == "프리미엄":
        mem = 7.0 * vc
        mem = min(mem, 560.0)
        return str(round(mem, 1))

    if hw == "프리미엄 메모리 최적화":
        mem = 13.6 * vc
        mem = min(mem, 870.4)
        return str(round(mem, 1))

    return ""


# =========================
# 5. Rows Builders
# =========================

def build_vm_rows(env: str, vms_basic: List[Dict[str, Any]], nic_ip_map: Dict[str, str], disk_map: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not vms_basic:
        return rows

    first_sub = vms_basic[0].get("subscriptionId")
    locations = [v.get("location") for v in vms_basic]
    size_map = build_vm_size_map(first_sub, locations) if first_sub else {}

    for v in vms_basic:
        vm_size = v.get("vmSize", "")
        spec = size_map.get(vm_size, {})

        os_parts = [p for p in [v.get("osType") or "", v.get("osOffer") or "", v.get("osSku") or ""] if p]
        os_str = " / ".join(os_parts)

        nic_id = (v.get("nicId") or "").lower()
        private_ip = nic_ip_map.get(nic_id, "")

        os_disk_id = (v.get("osDiskId") or "").lower()
        os_disk_info = disk_map.get(os_disk_id, {})
        os_disk_sku = os_disk_info.get("sku", "")
        os_disk_size = os_disk_info.get("size", "")

        data_disks = v.get("dataDisks")
        data_disk_ids: List[str] = []
        data_disk_names: List[str] = []

        if isinstance(data_disks, list):
            for dd in data_disks:
                if isinstance(dd, dict):
                    md = dd.get("managedDisk") or {}
                    did = (md.get("id") or "").lower()
                    dname = dd.get("name") or ""
                    if did:
                        data_disk_ids.append(did)
                    if dname:
                        data_disk_names.append(dname)

        labels: List[str] = []
        sizes: List[str] = []
        for idx, did in enumerate(data_disk_ids):
            info = disk_map.get(did, {})
            sku = info.get("sku", "")
            size = info.get("size", "")
            name = data_disk_names[idx] if idx < len(data_disk_names) else info.get("name", "")
            if name and sku:
                labels.append(f"{name}({sku})")
            elif name:
                labels.append(name)
            elif sku:
                labels.append(sku)
            if size != "" and size is not None:
                sizes.append(str(size))

        rows.append({
            "Environment": env,
            "SubscriptionID": v.get("subscriptionId",""),
            "ResourceGroup":  v.get("resourceGroup",""),
            "ResourceName":   v.get("name",""),
            "Type":           v.get("type",""),
            "Location":       v.get("location",""),
            "SKU":            vm_size,
            "vCPU":           spec.get("vCPU",""),
            "Memory":         spec.get("Memory",""),
            "OS":             os_str,
            "Private IP":     private_ip,
            "OS Disk":        os_disk_sku,
            "OS Disk Size(GiB)": os_disk_size,
            "Data Disk":      "; ".join(labels),
            "Data Disk Size(GiB)": "; ".join(sizes),
        })

    return rows

def build_vmss_rows(env: str, vmss_basic: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not vmss_basic:
        return rows

    first_sub = vmss_basic[0].get("subscriptionId")
    locations = [v.get("location") for v in vmss_basic]
    size_map = build_vm_size_map(first_sub, locations) if first_sub else {}

    for v in vmss_basic:
        vm_size = v.get("skuName", "")
        spec = size_map.get(vm_size, {})

        os_parts = [p for p in [v.get("osType") or "", v.get("osOffer") or "", v.get("osSku") or ""] if p]
        os_str = " / ".join(os_parts)

        # OS Disk
        os_disk_storage = v.get("osDiskStorageType", "") or ""
        os_disk_size = v.get("osDiskSizeGB", "")

        # Data Disks
        data_disks = v.get("dataDisks")
        labels: List[str] = []
        sizes: List[str] = []

        if isinstance(data_disks, list):
            for dd in data_disks:
                if not isinstance(dd, dict):
                    continue
                dname = dd.get("name") or ""
                dsize = dd.get("diskSizeGB", None)
                m = dd.get("managedDisk") or {}
                sku = m.get("storageAccountType", "") or ""
                if dname and sku:
                    labels.append(f"{dname}({sku})")
                elif dname:
                    labels.append(dname)
                elif sku:
                    labels.append(sku)
                if dsize not in (None, ""):
                    sizes.append(str(dsize))

        rows.append({
            "Environment": env,
            "SubscriptionID": v.get("subscriptionId",""),
            "ResourceGroup":  v.get("resourceGroup",""),
            "ResourceName":   v.get("name",""),
            "Type":           v.get("type",""),
            "Location":       v.get("location",""),
            "SKU":            vm_size,
            "vCPU":           spec.get("vCPU",""),
            "Memory":         spec.get("Memory",""),
            "OS":             os_str,
            "OS Disk":        os_disk_storage,
            "OS Disk Size(GiB)": os_disk_size,
            "Data Disk":      "; ".join(labels),
            "Data Disk Size(GiB)": "; ".join(sizes),
            "Orchestration Mode": v.get("orchestrationMode",""),
            "Instance Count": v.get("instanceCount",""),
        })

    return rows


def build_appgw_rows(appgws: List[Dict[str, Any]], env: str) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": ag.get("subscriptionId",""),
        "ResourceGroup":  ag.get("resourceGroup",""),
        "ResourceName":   ag.get("name",""),
        "Type":           ag.get("type",""),
        "Location":       ag.get("location",""),
        "SKU":            ag.get("skuType",""),
        "MinCapacity":    ag.get("minCapacity",""),
        "MaxCapacity":    ag.get("maxCapacity",""),
        "Public IP":      ag.get("publicIPAddress",""),
        "Private IP":     ag.get("privateIP",""),
    } for ag in appgws]

def build_acr_rows(acrs: List[Dict[str, Any]], env: str) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": a.get("subscriptionId",""),
        "ResourceGroup":  a.get("resourceGroup",""),
        "ResourceName":   a.get("name",""),
        "Type":           a.get("type",""),
        "Location":       a.get("location",""),
        "SKU":            a.get("skuName",""),
        "Login Server":   a.get("loginServer","")
    } for a in acrs]

def build_pg_rows(pgs: List[Dict[str, Any]], env: str) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": p.get("subscriptionId",""),
        "ResourceGroup":  p.get("resourceGroup",""),
        "ResourceName":   p.get("name",""),
        "Type":           p.get("type",""),
        "Location":       p.get("location",""),
        "SKU":            p.get("skuName",""),
        "Version":        p.get("version",""),
        "Storage Size":   p.get("storageSizeGB",""),
        "Storage IOPS":   p.get("storageIops",""),
        "HA":             p.get("HA",""),
        "Endpoint": p.get("fqdn","")
    } for p in pgs]

def build_kv_rows(kvs: List[Dict[str, Any]], env: str) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": kv.get("subscriptionId",""),
        "ResourceGroup":  kv.get("resourceGroup",""),
        "ResourceName":   kv.get("name",""),
        "Type":           kv.get("type",""),
        "Location":       kv.get("location",""),
        "SKU":            kv.get("sku",""),
        "Vault URI":      kv.get("vaultUri",""),
    } for kv in kvs]

def build_ca_rows(items: List[Dict[str, Any]], env: str, rg_name: str) -> List[Dict[str, Any]]:
    cache: Dict[Tuple[str, str], Dict[str, Dict[str, Any]]] = {}
    rows: List[Dict[str, Any]] = []

    for i in items:
        sub_id = i.get("subscriptionId", "")
        env_location = i.get("envLocation", "") or ""
        app_location = i.get("location", "") or ""

        wp_name = i.get("workloadProfileName", "") or ""
        wp_type = i.get("workloadProfileType", "") or ""

        cpu = ""
        mem = ""

        if sub_id and env_location and wp_type:
            key = (sub_id, env_location)
            if key not in cache:
                try:
                    cache[key] = fetch_available_workload_profile_types(sub_id, env_location, rg_name)
                except Exception as ex:
                    print(f"⚠ available workload profile 조회 실패 (sub={sub_id}, loc={env_location}): {ex}")
                    cache[key] = {}

            info = cache[key].get(wp_type, {})
            cpu = "" if info.get("cores") is None else str(info.get("cores"))
            mem = "" if info.get("memoryGiB") is None else str(info.get("memoryGiB"))

        rows.append({
            "Environment": env,
            "SubscriptionID": sub_id,
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       app_location,
            "Container App Environment": i.get("envName",""),
            "WorkloadProfile": wp_name,
            "WorkloadProfile CPU": cpu,
            "WorkloadProfile Memory": mem,
            "Application URL": i.get("appFqdn","")
        })
    return rows

def build_aks_rows_from_api(env: str, items: List[Dict[str, Any]], subscription_id: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not items:
        return rows

    locations = [i.get("location") for i in items if i.get("location")]
    size_map = build_vm_size_map(subscription_id, locations)

    for i in items:
        vm_size = i.get("vmSize", "")
        spec = size_map.get(vm_size, {})

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("clusterName",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "OS":             i.get("os",""),
            "Nodepool":       i.get("nodepoolName",""),
            "Node Count":     i.get("nodeCount",""),
            "Node CPU":       spec.get("vCPU",""),
            "Node Memory(GB)": spec.get("Memory",""),
            "OS Disk Size(GB)": i.get("osDiskSizeGB",""),
        })

    return rows

def build_storage_blob_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for s in items:
        sku = s.get("sku") or {}
        props = s.get("properties") or {}

        sku_name = ""
        sku_tier = ""
        if isinstance(sku, dict):
            sku_name = sku.get("name", "") or ""
            sku_tier = sku.get("tier", "") or ""
        else:
            sku_name = str(sku)

        redundancy = parse_redundancy(sku_name)

        hns = ""
        access_tier = ""
        if isinstance(props, dict):
            hns = str(props.get("isHnsEnabled", ""))
            access_tier = str(props.get("accessTier", ""))

        namespace = parse_namespace(hns)

        rows.append({
            "Environment": env,
            "SubscriptionID": s.get("subscriptionId",""),
            "ResourceGroup":  s.get("resourceGroup",""),
            "ResourceName":   s.get("name",""),
            "Type":           s.get("type",""),
            "Location":       s.get("location",""),
            "Kind":           s.get("kind",""),
            "Performance":    sku_tier,
            "Namespace":      namespace,
            "Access Tier":    access_tier,
            "Redundancy":     redundancy,
        })

    return rows


def build_file_storage_rows(
    env: str,
    storage_accounts: List[Dict[str, Any]],
    rg_name: str
) -> List[Dict[str, Any]]:

    rows: List[Dict[str, Any]] = []

    # ✅ 반드시 함수 안, for문 위에서 선언
    rg_target = (rg_name or "").strip().lower()

    for sa in storage_accounts:
        rg = sa.get("resourceGroup", "")
        if rg_target and rg.lower() != rg_target:
            continue   # 다른 RG면 스킵

        sub_id = sa.get("subscriptionId", "")
        account_name = sa.get("name", "")
        sa_location = sa.get("location", "")

        sku = sa.get("sku") or {}
        kind = sa.get("kind") or ""

        sku_name = ""
        sku_tier = ""
        if isinstance(sku, dict):
            sku_name = sku.get("name", "") or ""
            sku_tier = sku.get("tier", "") or ""
        else:
            sku_name = str(sku)

        redundancy = parse_redundancy_from_sku_name(sku_name)
        media_tier = parse_media_tier(sku_tier, kind, sku_name)

        try:
            shares = list_file_shares(sub_id, rg, account_name)
        except Exception as ex:
            print(f"⚠ File Share 목록 조회 실패: {sub_id}/{rg}/{account_name} / {ex}")
            continue

        for sh in shares:
            sh_name = sh.get("name", "")
            sh_type = sh.get("type", "") or "Microsoft.Storage/storageAccounts/fileServices/shares"

            props = sh.get("properties") or {}
            if not isinstance(props, dict):
                props = {}

            access_tier = map_access_tier(str(props.get("accessTier", "") or ""))
            billing = parse_billing_model(props, sku_tier, kind)

            quota_gb = ""
            q = props.get("shareQuota", None)
            if q is None:
                q = props.get("quota", None)
            if q is not None:
                try:
                    quota_gb = str(int(q))
                except Exception:
                    quota_gb = str(q)

            rows.append({
                "Environment": env,
                "SubscriptionID": sub_id,
                "ResourceGroup": rg,
                "ResourceName": sh_name,
                "Type": sh_type,
                "Location": sa_location,
                "Storage Account Name": account_name,
                "Storage Account Kind": kind,
                "Media Tier": media_tier,
                "Redundancy": redundancy,
                "Billing Model": billing,
                "Access Tier": access_tier,
                "Quota(GB)": quota_gb,
            })

    return rows

def build_aisearch_rows(items: List[Dict[str, Any]], env: str) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": s.get("subscriptionId",""),
        "ResourceGroup":  s.get("resourceGroup",""),
        "ResourceName":   s.get("name",""),
        "Type":           s.get("type",""),
        "Location":       s.get("location",""),
        "SKU":            s.get("skuName",""),
        "Replica Count":  s.get("replicaCount",""),
        "Partition Count": s.get("partitionCount",""),
        "URL":            s.get("url","")
    } for s in items]

def build_docint_rows(env: str, items: Any) -> List[Dict[str, Any]]:
    if not isinstance(items, list):
        print(f"⚠ build_docint_rows: items가 list가 아님 (type={type(items)}) -> 스킵")
        return []

    rows: List[Dict[str, Any]] = []
    for i in items:
        if not isinstance(i, dict):
            print(f"⚠ build_docint_rows: dict가 아닌 항목 발견 (type={type(i)}) -> 스킵: {i}")
            continue

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Kind":           i.get("kindName",""),
            "SKU":            i.get("skuName",""),
            "Endpoint":       i.get("endpoint","")
        })
    return rows

def build_aml_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Container Registry": i.get("acrName",""),
            "Key Vault":          i.get("kvName",""),
            "Application Insights": i.get("aiName",""),
        })
    return rows

def build_eh_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Tier":           i.get("tier",""),
            "Throughput Units": i.get("throughputUnits",""),
            "Auto Inflate":     i.get("autoInflate",""),
            "Max Throughput Units": i.get("maxThroughputUnits",""),
            "Premium Partitions": i.get("premiumPartitions",""),
            "SKU Capacity": i.get("skuCapacity",""),
        })
    return rows

def build_openai_rows(items: List[Dict[str, Any]], env: str) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": o.get("subscriptionId",""),
        "ResourceGroup":  o.get("resourceGroup",""),
        "ResourceName":   o.get("name",""),
        "Type":           o.get("type",""),
        "Location":       o.get("location",""),
        "SKU":            o.get("skuName",""),
        "Endpoint":       o.get("endpoint",""),
    } for o in items]

def build_mysql_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": i.get("subscriptionId",""),
        "ResourceGroup":  i.get("resourceGroup",""),
        "ResourceName":   i.get("name",""),
        "Type":           i.get("type",""),
        "Location":       i.get("location",""),
        "SKU":            i.get("skuName",""),
        "Version":        i.get("version",""),
        "Storage Size":   i.get("storageSizeGB",""),
        "Storage IOPS":   i.get("storageIops",""),
        "HA":             i.get("HA",""),
        "Endpoint":       i.get("fqdn","")
    } for i in items]

def build_managed_redis_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": i.get("subscriptionId",""),
        "ResourceGroup":  i.get("resourceGroup",""),
        "ResourceName":   i.get("name",""),
        "Type":           i.get("type",""),
        "Location":       i.get("location",""),
        "Tier":           i.get("tier",""),
        "Instance":       i.get("instance",""),
        "HA":             i.get("ha",""),
        "Endpoint": i.get("host","")
    } for i in items]

def build_docdb_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for i in items:
        sub_id = i.get("subscriptionId","")
        rg = i.get("resourceGroup","")
        name = i.get("name","")

        tier = shard = storage = ha = "N/A"

        if sub_id and rg and name:
            try:
                detail = fetch_documentdb_details(sub_id, rg, name)
                tier = detail.get("clusterTier") or "N/A"
                shard = detail.get("shardCount")
                storage = detail.get("storageGb")
                ha = detail.get("haMode") or "N/A"

                shard = "N/A" if shard in ("", None) else shard
                storage = "N/A" if storage in ("", None) else storage

            except Exception as ex:
                print(f"⚠ DocumentDB 상세 조회 실패: {sub_id}/{rg}/{name} / {ex}")

        rows.append({
            "Environment": env,
            "SubscriptionID": sub_id,
            "ResourceGroup":  rg,
            "ResourceName":   name,
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Cluster Tier":   tier,
            "Shard Count":    shard,
            "Storage Size(GB)": storage,
            "HA":             ha,
        })

    return rows

def build_cosmos_mongo_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for i in items:
        enable_multi_write = i.get("enableMultiWrite", False)

        locs = i.get("locs")
        write_regions = ""
        if isinstance(locs, list):
            names = []
            for x in locs:
                if isinstance(x, dict):
                    n = x.get("locationName")
                    if n:
                        names.append(str(n))
            write_regions = "; ".join(names)

        write_mode = "다중 쓰기 지역" if enable_multi_write else "단일 쓰기 지역"

        ttl = i.get("totalThroughputLimit", None)
        if ttl is None or ttl == "":
            ru_out = "N/A"
        elif isinstance(ttl, int) and ttl == -1:
            ru_out = "Unlimited (-1)"
        else:
            ru_out = str(ttl)

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Write Regions": write_regions,
            "Write Region Mode": write_mode,
            "RU Throughput": ru_out,
            "URI": i.get("uri","")
        })

    return rows

def build_cosmos_pg_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        node_count = i.get("nodeCount", None)

        if isinstance(node_count, int):
            tier = "단일 노드" if node_count == 0 else "다중 노드"
        else:
            tier = "N/A"

        coord = []
        if i.get("coordEdition"):
            coord.append(str(i.get("coordEdition")))
        if i.get("coordVCores") not in ("", None):
            coord.append(f"{i.get('coordVCores')} vCores")
        coord_str = " ".join(coord) if coord else "N/A"

        node = []
        if i.get("nodeEdition"):
            node.append(str(i.get("nodeEdition")))
        if i.get("nodeVCores") not in ("", None):
            node.append(f"{i.get('nodeVCores')} vCores")
        node_str = " ".join(node) if node else "N/A"

        if isinstance(node_count, int) and node_count >= 2:
            compute = f"Coordinator: {coord_str} / Worker: {node_str} (nodes={node_count})"
        elif isinstance(node_count, int) and node_count == 0:
            compute = f"Coordinator: {coord_str}"
        else:
            compute = f"Coordinator: {coord_str} / Worker: {node_str}"

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Tier":           tier,
            "Compute Size":   compute,
        })
    return rows

def build_sqlmi_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        sku_name = i.get("skuName","")
        svc_tier = parse_sqlmi_service_tier(sku_name)
        hw_gen = parse_sqlmi_hw_generation(sku_name)
        vcores = i.get("vcores","")
        mem_gb = estimate_sqlmi_memory_gb(hw_gen, vcores)

        zr = i.get("zoneRedundant", "")
        if zr is True:
            zr_out = "Yes"
        elif zr is False:
            zr_out = "No"
        else:
            zr_out = ""

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Service Tier":   svc_tier,
            "Hardware Generation": hw_gen,
            "vCore":          vcores if vcores is not None else "",
            "Memory(GB)":     mem_gb,
            "Storage Size(GB)": i.get("storageGB",""),
            "Zone Redundant": zr_out,
            "Host": i.get("host","")
        })
    return rows

def build_vnet_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        vnet_prefixes = i.get("vnetPrefixes")
        if isinstance(vnet_prefixes, list):
            vnet_ip_range = "; ".join([str(x) for x in vnet_prefixes if x])
        else:
            vnet_ip_range = str(vnet_prefixes or "")

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "vNet IP Range":  vnet_ip_range,
            "Subnet Name":    i.get("subnetName",""),
            "Subnet IP Range": i.get("subnetPrefix",""),
            "Subnet NSG":     i.get("nsgName",""),
            "Subnet UDR":     i.get("udrName",""),
        })
    return rows

def build_private_endpoint_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": i.get("subscriptionId",""),
        "ResourceGroup":  i.get("resourceGroup",""),
        "ResourceName":   i.get("name",""),
        "Type":           i.get("type",""),
        "Location":       i.get("location",""),
        "Private IP":     i.get("privateIP",""),   # ⭐ 조인 결과
        "Connected Resource": i.get("connected",""),
    } for i in items]



def build_law_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "SKU":            i.get("skuName",""),
            "Retention(Days)": i.get("retentionDays",""),
            "Public Network Access": i.get("publicNetworkAccess",""),
        })
    return rows


def build_appinsights_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Application Type": i.get("appType",""),
            "Workspace(ResourceId)": i.get("workspaceResourceId",""),
        })
    return rows

def build_functionapp_rows(env: str, items: List[Dict[str, Any]], plan_map: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        plan_id = (i.get("planId") or "").lower()
        plan_name = ""
        plan_tier = ""
        if plan_id and plan_id in plan_map:
            plan_name = plan_map[plan_id].get("name","")
            plan_tier = plan_map[plan_id].get("pricingTier","")

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Default Domain": i.get("defaultHost",""),
            "OS":             i.get("os",""),
            "App Service Plan": plan_name,
            "Pricing Tier":     plan_tier,
            "Runtime Version":  i.get("runtime",""),
        })

    return rows

def build_logicapp_rows(env: str, items: List[Dict[str, Any]], plan_map: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        plan_id = (i.get("planId") or "").lower()

        # Consumption은 planId가 없어서 N/A 처리
        plan_name = "N/A"
        plan_tier = "N/A"
        if plan_id and plan_id in plan_map:
            plan_name = plan_map[plan_id].get("name","")
            plan_tier = plan_map[plan_id].get("pricingTier","")

        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "Default Domain": i.get("defaultHost",""),
            "OS":             i.get("os",""),
            "App Service Plan": plan_name,
            "Pricing Tier":     plan_tier,
            "Runtime Version":  i.get("runtime",""),
        })

    return rows

def _join_list(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return "; ".join([str(x) for x in v if x not in (None, "", "null")])
    return str(v)

def build_lb_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for i in items:
        rows.append({
            "Environment": env,
            "SubscriptionID": i.get("subscriptionId",""),
            "ResourceGroup":  i.get("resourceGroup",""),
            "ResourceName":   i.get("name",""),
            "Type":           i.get("type",""),
            "Location":       i.get("location",""),
            "SKU":            i.get("SKU",""),
            "Tier":           i.get("Tier",""),
            "Frontend IP Name": _join_list(i.get("FrontendNames")),
            "Frontend IP":      _join_list(i.get("FrontendIPs")),
            "Backend Pool Name": _join_list(i.get("BackendPoolNames")),
            "Load Balancing Rule Name": _join_list(i.get("LBRuleNames")),
            "Health Probe Name": _join_list(i.get("ProbeNames")),
        })
    return rows

def build_managed_identity_rows(env: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [{
        "Environment": env,
        "SubscriptionID": i.get("subscriptionId",""),
        "ResourceGroup":  i.get("resourceGroup",""),
        "ResourceName":   i.get("name",""),
        "Type":           i.get("type",""),
        "Location":       i.get("location",""),
    } for i in items]


# =========================
# 6. Excel 저장 (시트별)
# =========================

def write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

def autosize_columns(ws, headers: List[str], max_width: int = 60) -> None:
    for i, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for col_cells in ws.iter_cols(min_col=i, max_col=i, min_row=2, max_row=min(ws.max_row, 200)):
            for cell in col_cells:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, max_width)

def ensure_sheet(wb, sheet_name: str, headers: List[str]):
    sname = sheet_name[:31]
    if sname in wb.sheetnames:
        ws = wb[sname]
        # 헤더가 비어있으면 넣기
        if ws.max_row == 0:
            ws.append(headers)
        elif ws.max_row == 1 and ws.cell(1, 1).value in (None, ""):
            ws.delete_rows(1, 1)
            ws.append(headers)
        return ws

    ws = wb.create_sheet(title=sname)
    ws.append(headers)
    return ws

def append_rows(ws, headers: List[str], rows: List[Dict[str, Any]]):
    if not rows:
        return
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

def autosize_columns_safe(ws, headers: List[str], max_width: int = 60):
    # 기존 autosize는 append마다 전체 스캔하면 느려져서,
    # 새로 추가된 200행 정도만 보는 방식으로 가볍게
    max_row = ws.max_row
    start_row = max(2, max_row - 200)

    for i, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in range(start_row, max_row + 1):
            v = ws.cell(row=row, column=i).value
            v = "" if v is None else str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, max_width)

def write_xlsx(path: str, sheets: Dict[str, Dict[str, Any]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, payload in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        headers = payload["headers"]
        rows = payload["rows"]
        write_sheet(ws, headers, rows)
        autosize_columns(ws, headers)

    wb.save(path)



# =========================
# 7. main
# =========================

def main():
    today = datetime.now().strftime("%Y%m%d")
    xlsx_filename = f"{service_name}_{service_abbr}_inventory_{today}.xlsx"
    xlsx_path = os.path.join(out_dir, xlsx_filename)

    # =========================
    # 1) 결과 누적 리스트
    # =========================
    all_appgw_rows: List[Dict[str, Any]] = []
    all_lb_rows: List[Dict[str, Any]] = []
    all_vm_rows: List[Dict[str, Any]] = []
    all_vmss_rows: List[Dict[str, Any]] = []
    all_acr_rows: List[Dict[str, Any]] = []
    all_ca_rows: List[Dict[str, Any]] = []

    all_functionapp_rows: List[Dict[str, Any]] = []
    all_logicapp_rows: List[Dict[str, Any]] = []
    all_appservice_rows: List[Dict[str, Any]] = []   # (코드에 함수가 있으면 채움)

    all_aks_rows: List[Dict[str, Any]] = []
    all_kv_rows: List[Dict[str, Any]] = []

    all_st_blob_rows: List[Dict[str, Any]] = []
    all_file_rows: List[Dict[str, Any]] = []

    all_aisearch_rows: List[Dict[str, Any]] = []
    all_docint_rows: List[Dict[str, Any]] = []
    all_aml_rows: List[Dict[str, Any]] = []
    all_eh_rows: List[Dict[str, Any]] = []
    all_openai_rows: List[Dict[str, Any]] = []

    all_pg_rows: List[Dict[str, Any]] = []
    all_mysql_rows: List[Dict[str, Any]] = []
    all_managed_redis_rows: List[Dict[str, Any]] = []

    all_cosmos_rows: List[Dict[str, Any]] = []       # (코드에 Cosmos 전체 함수가 있으면 채움)
    all_cosmos_mongo_rows: List[Dict[str, Any]] = []
    all_cosmos_pg_rows: List[Dict[str, Any]] = []

    all_docdb_rows: List[Dict[str, Any]] = []
    all_sqlmi_rows: List[Dict[str, Any]] = []

    all_vnet_rows: List[Dict[str, Any]] = []
    all_private_endpoint_rows: List[Dict[str, Any]] = []

    all_law_rows: List[Dict[str, Any]] = []
    all_appinsights_rows: List[Dict[str, Any]] = []

    all_managed_identity_rows: List[Dict[str, Any]] = []

    # =========================
    # 2) 타입 게이트(스캔 기반)
    # =========================
    T_APPGW = ["microsoft.network/applicationgateways"]
    T_LB = ["microsoft.network/loadbalancers"]

    T_VM = ["microsoft.compute/virtualmachines"]
    T_VMSS = ["microsoft.compute/virtualmachinescalesets"]
    T_NIC = ["microsoft.network/networkinterfaces"]
    T_DISK = ["microsoft.compute/disks"]

    T_ACR = ["microsoft.containerregistry/registries"]
    T_CA = ["microsoft.app/containerapps", "microsoft.app/managedenvironments"]

    T_WEB_SITES = ["microsoft.web/sites"]                 # Function/Logic(Standard)/AppService
    T_LOGIC_CONS = ["microsoft.logic/workflows"]          # Logic Consumption
    T_SERVERFARMS = ["microsoft.web/serverfarms"]         # App Service Plan

    T_AKS = ["microsoft.containerservice/managedclusters"]

    T_STORAGE = ["microsoft.storage/storageaccounts"]
    T_AISEARCH = ["microsoft.search/searchservices"]
    T_COG = ["microsoft.cognitiveservices/accounts"]      # Doc Intelligence + OpenAI
    T_AML = ["microsoft.machinelearningservices/workspaces"]
    T_EH = ["microsoft.eventhub/namespaces"]

    T_PG = ["microsoft.dbforpostgresql/flexibleservers", "microsoft.dbforpostgresql/servers"]
    T_MYSQL = ["microsoft.dbformysql/flexibleservers"]
    T_REDIS = ["microsoft.cache/redisenterprise"]

    T_COSMOS_ACCT = ["microsoft.documentdb/databaseaccounts"]  # Cosmos DB (Core/SQL/MongoDB API(RU))
    T_COSMOS_PG = ["microsoft.dbforpostgresql/servergroupsv2"] # CosmosDB for PostgreSQL

    T_DOCDB = ["microsoft.documentdb/mongoclusters"]           # Document DB (MongoDB vCore)
    T_SQLMI = ["microsoft.sql/managedinstances"]

    T_VNET = ["microsoft.network/virtualnetworks"]
    T_PE = ["microsoft.network/privateendpoints"]

    T_KV = ["microsoft.keyvault/vaults"]
    T_LAW = ["microsoft.operationalinsights/workspaces"]
    T_APPINSIGHTS = ["microsoft.insights/components"]

    T_MI = ["microsoft.managedidentity/userassignedidentities"]

    # =========================
    # 3) sheets: rows 있는 것만 만들기 + 타입 존재하는 것만 시트 생성
    # =========================
    def add_sheet_if_rows(sheets: Dict[str, Dict[str, Any]], sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]):
        if rows and len(rows) > 0:
            sheets[sheet_name] = {"headers": headers, "rows": rows}

    # =========================
    # 4) env / rg 반복 (중복 실행 방지)
    # =========================
    processed_keys: Set[Tuple[str, str, str]] = set()  # (env, sub_id, rg_lower)

    # 시트 생성 여부를 "실제 타입 존재"로 제어하기 위한 플래그들
    present = {
        "appgw": False, "lb": False, "vm": False, "vmss": False, "acr": False, "ca": False,
        "function": False, "logic": False, "aks": False,
        "storage": False, "aisearch": False, "docint": False, "openai": False,
        "aml": False, "eh": False, "pg": False, "mysql": False, "redis": False,
        "cosmos_acct": False, "cosmos_pg": False, "docdb": False, "sqlmi": False,
        "vnet": False, "pe": False, "kv": False, "law": False, "appinsights": False,
        "serverfarms": False, "websites": False, "logic_cons": False, "managed_identity": False
    }

    for env, sub_id in SUB_ENV_MAP.items():
        rg_list = RG_ENV_MAP.get(env, [])
        if not rg_list:
            print(f"\n========== {env} ({sub_id}) ==========")
            print("⚠ Resource Group이 입력되지 않아 이 환경은 스킵합니다.")
            continue

        for rg_name in rg_list:
            rg_name = (rg_name or "").strip()
            if not rg_name:
                continue

            key = (env, sub_id, rg_name.lower())
            if key in processed_keys:
                print(f"⚠ 중복 RG 감지로 스킵: {env}/{sub_id}/{rg_name}")
                continue
            processed_keys.add(key)

            print(f"\n========== {env} ({sub_id}) / RG={rg_name} ==========")

            # (1) RG 범위 타입 스캔 (RG당 정확히 1번)
            try:
                types_in_rg = discover_resource_types(sub_id, rg_name)
                scan_failed = False
                print(f"👉 리소스 타입 스캔 완료(RG 범위): {len(types_in_rg)} types")
            except Exception as ex:
                print(f"⚠ 타입 스캔 실패 (sub={sub_id}, rg={rg_name}) -> 전체 쿼리 실행 모드: {ex}")
                types_in_rg = set()
                scan_failed = True

            # ✅ present 플래그(“실제 배포 여부”) 업데이트
            def mark(flag: str, expected: List[str]):
                if scan_failed:
                    # 스캔 실패 시엔 "실제 존재" 판단을 못하므로, 시트 생성은 rows 기준으로만 가게 된다
                    return
                if should_run(types_in_rg, expected):
                    present[flag] = True

            mark("appgw", T_APPGW)
            mark("lb", T_LB)
            mark("vm", T_VM)
            mark("acr", T_ACR)
            mark("ca", T_CA)
            mark("websites", T_WEB_SITES)
            mark("logic_cons", T_LOGIC_CONS)
            mark("serverfarms", T_SERVERFARMS)
            mark("aks", T_AKS)
            mark("storage", T_STORAGE)
            mark("aisearch", T_AISEARCH)
            mark("aml", T_AML)
            mark("eh", T_EH)
            mark("pg", T_PG)
            mark("mysql", T_MYSQL)
            mark("redis", T_REDIS)
            mark("cosmos_acct", T_COSMOS_ACCT)
            mark("cosmos_pg", T_COSMOS_PG)
            mark("docdb", T_DOCDB)
            mark("sqlmi", T_SQLMI)
            mark("vnet", T_VNET)
            mark("pe", T_PE)
            mark("kv", T_KV)
            mark("law", T_LAW)
            mark("appinsights", T_APPINSIGHTS)
            # DocInt/OpenAI는 같은 계열(COG)이라 따로 표기
            mark("docint", T_COG)
            mark("openai", T_COG)
            mark("managed_identity", T_MI)

            # VM
            if scan_failed or should_run(types_in_rg, T_VM):
                vms_basic = query_azure_vms_basic([sub_id], rg_name)
                if vms_basic:
                    nic_ip_map = query_nic_private_ips([sub_id], rg_name) if (scan_failed or should_run(types_in_rg, T_NIC)) else {}
                    disk_map = query_disks([sub_id], rg_name) if (scan_failed or should_run(types_in_rg, T_DISK)) else {}
                    all_vm_rows.extend(build_vm_rows(env, vms_basic, nic_ip_map, disk_map))
            
            # VMSS
            if scan_failed or should_run(types_in_rg, T_VMSS):
                vmss_basic = query_azure_vmss_basic([sub_id], rg_name)
                if vmss_basic:
                    all_vmss_rows.extend(build_vmss_rows(env, vmss_basic))


            # Application Gateway
            if scan_failed or should_run(types_in_rg, T_APPGW):
                appgws = query_azure_appgws([sub_id], rg_name)
                if appgws:
                    all_appgw_rows.extend(build_appgw_rows(appgws, env))

            # Load Balancer (REST)
            if scan_failed or should_run(types_in_rg, T_LB):
                lbs = query_load_balancers_rest(sub_id, rg_name)
                if lbs:
                    all_lb_rows.extend(build_lb_rows(env, lbs))

            # ACR
            if scan_failed or should_run(types_in_rg, T_ACR):
                acrs = query_azure_acrs([sub_id], rg_name)
                if acrs:
                    all_acr_rows.extend(build_acr_rows(acrs, env))

            # Container Apps
            if scan_failed or should_run(types_in_rg, T_CA):
                cas = query_azure_containerapps_with_workload_profiles([sub_id], rg_name)
                if cas:
                    all_ca_rows.extend(build_ca_rows(cas, env, rg_name))

            # App Service Plan map (Function/Logic/AppService에서 공용)
            plan_map: Dict[str, Dict[str, str]] = {}
            if scan_failed or should_run(types_in_rg, T_SERVERFARMS):
                try:
                    plan_map = query_app_service_plans([sub_id], rg_name)
                except Exception as ex:
                    print(f"⚠ App Service Plan 매핑 실패: {ex}")
                    plan_map = {}

            # Function Apps
            if scan_failed or should_run(types_in_rg, T_WEB_SITES):
                fas = query_function_apps([sub_id], rg_name)
                if fas:
                    present["function"] = True
                    all_functionapp_rows.extend(build_functionapp_rows(env, fas, plan_map))

            # Logic Apps (Standard + Consumption)
            if scan_failed or should_run(types_in_rg, (T_WEB_SITES + T_LOGIC_CONS)):
                las = query_logic_apps([sub_id], rg_name)
                if las:
                    present["logic"] = True
                    all_logicapp_rows.extend(build_logicapp_rows(env, las, plan_map))

            # App Service (현재 코드에는 query/build/header가 없다 했으니, 함수가 생기면 자동으로)
            # q_as = globals().get("query_app_services")
            # b_as = globals().get("build_appservice_rows")
            # h_as = globals().get("appservice_headers")
            # if q_as and b_as and h_as and (scan_failed or should_run(types_in_rg, T_WEB_SITES)):
            #     try:
            #         appsrv = q_as([sub_id], rg_name)
            #         if appsrv:
            #             all_appservice_rows.extend(b_as(env, appsrv, plan_map))
            #     except Exception as ex:
            #         print(f"⚠ App Service 조회 실패: {ex}")
            # elif not (q_as and b_as and h_as):
            #     # 기존 출력 유지
            #     print("ℹ App Service: query/build/header 함수가 코드에 없어 스킵합니다. (필요 시 추가 가능)")

            # AKS (SDK)
            if scan_failed or should_run(types_in_rg, T_AKS):
                aks_items = query_aks_nodepools_api(sub_id, rg_name)
                if aks_items:   
                    all_aks_rows.extend(build_aks_rows_from_api(env, aks_items, sub_id))

            # Storage Account + Azure Files
            if scan_failed or should_run(types_in_rg, T_STORAGE):
                sas = query_storage_accounts_blob_config([sub_id], rg_name)
                if sas:
                    all_st_blob_rows.extend(build_storage_blob_rows(env, sas))
                    file_rows = build_file_storage_rows(env, sas, rg_name)
                    if file_rows:
                        all_file_rows.extend(file_rows)

            # AI Search
            if scan_failed or should_run(types_in_rg, T_AISEARCH):
                ais = query_azure_ai_search([sub_id], rg_name)
                if ais:
                    all_aisearch_rows.extend(build_aisearch_rows(ais, env))

            # Document Intelligence
            if scan_failed or should_run(types_in_rg, T_COG):
                docints = query_azure_document_intelligence([sub_id], rg_name)
                if docints:
                    all_docint_rows.extend(build_docint_rows(env, docints))

            # AML
            if scan_failed or should_run(types_in_rg, T_AML):
                amls = query_azure_machine_learning_workspaces([sub_id], rg_name)
                if amls:
                    all_aml_rows.extend(build_aml_rows(env, amls))

            # Event Hubs
            if scan_failed or should_run(types_in_rg, T_EH):
                ehs = query_azure_eventhub_namespaces([sub_id], rg_name)
                if ehs:
                    all_eh_rows.extend(build_eh_rows(env, ehs))

            # Azure OpenAI
            if scan_failed or should_run(types_in_rg, T_COG):
                oais = query_azure_openai([sub_id], rg_name)
                if oais:
                    all_openai_rows.extend(build_openai_rows(oais, env))

            # PostgreSQL
            if scan_failed or should_run(types_in_rg, T_PG):
                pgs = query_azure_postgresql([sub_id], rg_name)
                if pgs:
                    all_pg_rows.extend(build_pg_rows(pgs, env))

            # MySQL
            if scan_failed or should_run(types_in_rg, T_MYSQL):
                mysqls = query_azure_mysql_flexible([sub_id], rg_name)
                if mysqls:
                    all_mysql_rows.extend(build_mysql_rows(env, mysqls))

            # Managed Redis
            if scan_failed or should_run(types_in_rg, T_REDIS):
                reds = query_azure_managed_redis([sub_id], rg_name)
                if reds:
                    all_managed_redis_rows.extend(build_managed_redis_rows(env, reds))

            # Cosmos DB (전체) - 함수가 있을 때만
            # q_cos_all = globals().get("query_cosmosdb_accounts_all")
            # b_cos_all = globals().get("build_cosmos_rows")
            # h_cos_all = globals().get("cosmos_headers")
            # if q_cos_all and b_cos_all and h_cos_all and (scan_failed or should_run(types_in_rg, T_COSMOS_ACCT)):
            #     try:
            #         cos_all = q_cos_all([sub_id], rg_name)
            #         if cos_all:
            #             all_cosmos_rows.extend(b_cos_all(env, cos_all))
            #     except Exception as ex:
            #         print(f"⚠ CosmosDB(전체) 조회 실패(함수 존재하나 실행 실패): {ex}")
            # elif (scan_failed or should_run(types_in_rg, T_COSMOS_ACCT)) and not (q_cos_all and b_cos_all and h_cos_all):
            #     print("ℹ CosmosDB(전체): query/build/header 함수가 코드에 없어 스킵합니다. (Mongo/PostgreSQL은 별도 탭으로 추출 중)")

            # CosmosDB for Mongo API (RU)
            if scan_failed or should_run(types_in_rg, T_COSMOS_ACCT):
                cosmos_mongo = query_cosmosdb_mongo_accounts([sub_id], rg_name)
                if cosmos_mongo:
                    all_cosmos_mongo_rows.extend(build_cosmos_mongo_rows(env, cosmos_mongo))

            # CosmosDB for PostgreSQL
            if scan_failed or should_run(types_in_rg, T_COSMOS_PG):
                cpg = query_cosmosdb_postgresql([sub_id], rg_name)
                if cpg:
                    all_cosmos_pg_rows.extend(build_cosmos_pg_rows(env, cpg))

            # Azure Document DB (Mongo vCore)
            if scan_failed or should_run(types_in_rg, T_DOCDB):
                docdb_list = query_azure_documentdb_list([sub_id], rg_name)
                if docdb_list:
                    all_docdb_rows.extend(build_docdb_rows(env, docdb_list))

            # SQL MI
            if scan_failed or should_run(types_in_rg, T_SQLMI):
                sqlmis = query_azure_sql_managed_instances([sub_id], rg_name)
                if sqlmis:
                    all_sqlmi_rows.extend(build_sqlmi_rows(env, sqlmis))

            # Key Vault
            if scan_failed or should_run(types_in_rg, T_KV):
                kvs = query_azure_keyvaults([sub_id], rg_name)
                if kvs:
                    all_kv_rows.extend(build_kv_rows(kvs, env))

            # VNet/Subnet
            if scan_failed or should_run(types_in_rg, T_VNET):
                vnets = query_virtual_networks_with_subnets([sub_id], rg_name)
                if vnets:
                    all_vnet_rows.extend(build_vnet_rows(env, vnets))

            # Private Endpoint
            if scan_failed or should_run(types_in_rg, T_PE):
                pes = query_private_endpoints([sub_id], rg_name)
                if pes:
                    all_private_endpoint_rows.extend(build_private_endpoint_rows(env, pes))

            # Log Analytics Workspace
            if scan_failed or should_run(types_in_rg, T_LAW):
                laws = query_log_analytics_workspaces([sub_id], rg_name)
                if laws:
                    all_law_rows.extend(build_law_rows(env, laws))

            # Application Insights
            if scan_failed or should_run(types_in_rg, T_APPINSIGHTS):
                apps = query_application_insights([sub_id], rg_name)
                if apps:
                    all_appinsights_rows.extend(build_appinsights_rows(env, apps))
            
            # Managed Identity (User Assigned)
            if scan_failed or should_run(types_in_rg, T_MI):
                mis = query_user_assigned_managed_identities([sub_id], rg_name)
                if mis:
                    all_managed_identity_rows.extend(build_managed_identity_rows(env, mis))


    # =========================
    # 5) Excel 저장: "실제 존재하는 타입" + "rows 있는 시트"만 생성
    #    - 스캔 실패한 경우(scan_failed=True)에는 present 판단이 어려우므로 rows 기준으로만 생성됨
    # =========================
    sheets: Dict[str, Dict[str, Any]] = {}

    # --- Summary (엑셀 첫 시트) ---
    sub_name_map = get_subscription_name_map(list(SUB_ENV_MAP.values()))
    summary_rows = build_summary_rows(
        all_rows_by_category=[
            all_appgw_rows, all_lb_rows, all_vm_rows, all_vmss_rows,
            all_acr_rows, all_ca_rows,
            all_functionapp_rows, all_logicapp_rows, all_appservice_rows,
            all_aks_rows, all_kv_rows,
            all_st_blob_rows, all_file_rows,
            all_aisearch_rows, all_docint_rows, all_aml_rows, all_eh_rows, all_openai_rows,
            all_pg_rows, all_mysql_rows, all_managed_redis_rows,
            all_cosmos_rows, all_cosmos_mongo_rows, all_cosmos_pg_rows,
            all_docdb_rows, all_sqlmi_rows,
            all_vnet_rows, all_private_endpoint_rows,
            all_law_rows, all_appinsights_rows
        ],
        sub_name_map=sub_name_map
    )
    
    # write_xlsx는 dict 삽입 순서대로 시트를 생성하므로, Summary를 제일 먼저 넣으면 "첫 시트"가 됨
    sheets["Summary"] = {"headers": summary_headers, "rows": summary_rows}


    # Networking / Compute
    if present["appgw"] or all_appgw_rows:
        add_sheet_if_rows(sheets, "Application Gateway", appgw_headers, all_appgw_rows)
    if present["lb"] or all_lb_rows:
        add_sheet_if_rows(sheets, "Load Balancer", lb_headers, all_lb_rows)
    if present["vm"] or all_vm_rows:
        add_sheet_if_rows(sheets, "Virtual Machine", vm_headers, all_vm_rows)
    if present["vmss"] or all_vmss_rows:
        add_sheet_if_rows(sheets, "VM Scale Set", vmss_headers, all_vmss_rows)

    # Container
    if present["acr"] or all_acr_rows:
        add_sheet_if_rows(sheets, "Container Registry", acr_headers, all_acr_rows)
    if present["ca"] or all_ca_rows:
        add_sheet_if_rows(sheets, "Container Apps", ca_headers, all_ca_rows)

    # App Platform
    if present["function"] or all_functionapp_rows:
        add_sheet_if_rows(sheets, "Function App", functionapp_headers, all_functionapp_rows)
    if present["logic"] or all_logicapp_rows:
        add_sheet_if_rows(sheets, "Logic App", logicapp_headers, all_logicapp_rows)

    if all_appservice_rows and globals().get("appservice_headers"):
        add_sheet_if_rows(sheets, "App Service", globals()["appservice_headers"], all_appservice_rows)

    if present["aks"] or all_aks_rows:
        add_sheet_if_rows(sheets, "AKS", aks_headers, all_aks_rows)

    # Storage
    if present["storage"] or all_st_blob_rows:
        add_sheet_if_rows(sheets, "Storage Account", st_sa_headers, all_st_blob_rows)
        add_sheet_if_rows(sheets, "File Storage", file_headers, all_file_rows)

    # AI
    if present["aisearch"] or all_aisearch_rows:
        add_sheet_if_rows(sheets, "AI Search", aisearch_headers, all_aisearch_rows)
    # COG가 존재해도 FormRecognizer/OpenAI가 없을 수 있어서 rows로 판단
    add_sheet_if_rows(sheets, "Document Intelligence", docint_headers, all_docint_rows)
    add_sheet_if_rows(sheets, "Azure OpenAI", openai_headers, all_openai_rows)

    if present["aml"] or all_aml_rows:
        add_sheet_if_rows(sheets, "Azure Machine Learning", aml_headers, all_aml_rows)
    if present["eh"] or all_eh_rows:
        add_sheet_if_rows(sheets, "Event Hubs", eh_headers, all_eh_rows)

    # DB
    if present["pg"] or all_pg_rows:
        add_sheet_if_rows(sheets, "PostgreSQL", pg_headers, all_pg_rows)
    if present["mysql"] or all_mysql_rows:
        add_sheet_if_rows(sheets, "MySQL", mysql_headers, all_mysql_rows)
    if present["redis"] or all_managed_redis_rows:
        add_sheet_if_rows(sheets, "Azure Managed Redis", managed_redis_headers, all_managed_redis_rows)

    # Cosmos
    if all_cosmos_rows and globals().get("cosmos_headers"):
        add_sheet_if_rows(sheets, "Azure Cosmos DB", globals()["cosmos_headers"], all_cosmos_rows)
    if present["cosmos_acct"] or all_cosmos_mongo_rows:
        add_sheet_if_rows(sheets, "CosmosDB for MongoDB", cosmos_mongo_headers, all_cosmos_mongo_rows)
    if present["cosmos_pg"] or all_cosmos_pg_rows:
        add_sheet_if_rows(sheets, "CosmosDB for PostgreSQL", cosmos_pg_headers, all_cosmos_pg_rows)

    if present["docdb"] or all_docdb_rows:
        add_sheet_if_rows(sheets, "Azure DocumentDB", docdb_headers, all_docdb_rows)
    if present["sqlmi"] or all_sqlmi_rows:
        add_sheet_if_rows(sheets, "SQL Managed Instance", sqlmi_headers, all_sqlmi_rows)

    # Network / Ops
    if present["kv"] or all_kv_rows:
        add_sheet_if_rows(sheets, "Key Vault", kv_headers, all_kv_rows)
    if present["vnet"] or all_vnet_rows:
        add_sheet_if_rows(sheets, "Virtual Network", vnet_headers, all_vnet_rows)
    if present["pe"] or all_private_endpoint_rows:
        add_sheet_if_rows(sheets, "Private Endpoint", private_endpoint_headers, all_private_endpoint_rows)

    if present["law"] or all_law_rows:
        add_sheet_if_rows(sheets, "Log Analytics", law_headers, all_law_rows)
    if present["appinsights"] or all_appinsights_rows:
        add_sheet_if_rows(sheets, "Application Insights", appinsights_headers, all_appinsights_rows)

    if present.get("managed_identity") or all_managed_identity_rows:
        add_sheet_if_rows(sheets, "Managed Identity", managed_identity_headers, all_managed_identity_rows)

    if not sheets:
        raise RuntimeError("추출된 리소스가 없습니다. (권한/구독ID/RG 입력값을 확인해주세요)")

    write_xlsx(xlsx_path, sheets)

    print("\n✅ Excel 생성 완료:")
    print(" -", xlsx_path)


if __name__ == "__main__":
    main()
